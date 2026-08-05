#!/usr/bin/env python3
"""
生成分支报表脚本 - 第二步到第四步 (v3.5)
用法: python3 generate_branch.py <主报表.xlsx> <车辆进出报表.xlsx> <输出目录>

v3.5: 统一页眉 &L&G (左上角 Logo) + 标准化页面设置 (A4 / 水平居中 / 统一页边距)
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintOptions
from openpyxl.drawing.image import Image as _XLImage
import os, sys, re
from datetime import datetime, date, timedelta
from copy import copy
import math
import zipfile
import io
import shutil


# ============================================================
# Config
# ============================================================
MAIN_REPORT = None
VEHICLE_REPORT = None
OUTPUT_DIR = None

# ============================================================
# Merchant billing file definitions
# ============================================================
MERCHANT_FILES = {
    '爱康':     {'filename': '爱康国宾10元.xlsx',
                  'title': '爱 康 国 宾 停 车 券 计 费 明 细 表', 'price': 10},
    '爱康口腔': {'filename': '爱康口腔16元.xlsx',
                  'title': '爱 康 口腔 停 车 券 计 费 明 细 表', 'price': 16},
    '爱来':     {'filename': '爱来整形美容5元.xlsx',
                  'title': '爱 来 美 容 院 停 车 券 计 费 明 细 表', 'price': 5},
    '阿玛施':   {'filename': '阿玛施眼科电子券5元.xlsx',
                  'title': '眼 科 停 车 券 计 费 明 细 表', 'price': 5},
    '歌蕊':     {'filename': '歌蕊医疗13元.xlsx',
                  'title': '歌蕊医疗停车券计费明细表', 'price': 13},
    '皆大欢洗':  {'filename': '李利军3元.xlsx',
                  'title': '李 利 军 停 车 券 计 费 明 细 表', 'price': 3},
    '新奥美嘉':  {'filename': '新奥美嘉10元.xlsx',
                  'title': '新奥美嘉 停 车 券 计 费 明 细 表', 'price': 10},
    '丝芭':     {'filename': '丝芭16元.xlsx',
                  'title': '丝芭停 车 券 计 费 明 细 表', 'price': 16},
    '精应':     {'filename': '精应电子券4元.xlsx',
                  'title': '精 应 停 车 券 计 费 明 细 表', 'price': 4},
}

# ============================================================
# Standardized page setup (v3.5 — matching &L&G header from 李利军3元.xls)
# ============================================================
# Page margins in inches (v3.5 — tight margins for single-page fit with wide rows)
_PM_LEFT   = 0.4     # left margin
_PM_RIGHT  = 0.4     # right margin
_PM_TOP    = 0.5     # top margin
_PM_BOTTOM = 0.5     # bottom margin
_PAPER_A4  = 9       # openpyxl paper size code

def _apply_page_setup(ws):
    """Apply standardized page setup to a worksheet (v3.5).

    Matches the &L&G header layout from the reference 李利军3元.xls:
    - A4 paper, portrait
    - Fit to 1 page wide × 1 page tall (pageSetUpPr fitToPage="1" → Excel: 将工作表打印在一页)
    - Horizontal centering ON, vertical OFF
    - Standardized margins
    - Header logo via openpyxl oddHeader + VML legacyDrawingHF
    """
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = _PAPER_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options = PrintOptions(horizontalCentered=True, verticalCentered=False)
    ws.page_margins.left = _PM_LEFT
    ws.page_margins.right = _PM_RIGHT
    ws.page_margins.top = _PM_TOP
    ws.page_margins.bottom = _PM_BOTTOM
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    # Page header logo: floating image at A1 (compatible with Excel/WPS/Numbers)
    _add_logo_image(ws)

# ============================================================
# Date helpers
# ============================================================
def excel_serial_to_date(serial):
    """Convert Excel serial number / datetime / date / 'YYYY-MM-DD' string to date/datetime."""
    if serial is None:
        return None
    if isinstance(serial, datetime):
        return serial
    if isinstance(serial, date):
        return serial
    if isinstance(serial, (int, float)):
        # Excel epoch: 1899-12-30 (day 0)
        base = datetime(1899, 12, 30)
        try:
            return base + timedelta(days=int(serial))
        except:
            return None
    if isinstance(serial, str):
        m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', serial.strip())
        if m:
            try:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                return None
    return None

def parse_date_value(v):
    """Parse datetime / date / Excel serial / 'YYYY-MM-DD' string to a date object (返回 date)."""
    d = excel_serial_to_date(v)
    if d is None:
        return None
    return d.date() if isinstance(d, datetime) else d

# ============================================================
# Styles for 集团/九号 & 车场报表 (等价原版 HTML)
# ============================================================
from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side

def _mk_border(top='thin', bottom='thin', left='thin', right='thin',
               tc='B0B0B0', bc='B0B0B0', lc='B0B0B0', rc='B0B0B0'):
    def _s(st, c):
        return _Side(style=st, color=c) if st else _Side()
    return _Border(top=_s(top, tc), bottom=_s(bottom, bc),
                   left=_s(left, lc), right=_s(right, rc))

_B_THIN_J    = _mk_border()
_B_MED_TOP_J = _mk_border(top='medium', tc='333333')
_SJ_HDR  = {'font': _Font(name='微软雅黑', size=10, bold=True, color='1A4472'),
            'fill': _Fill('solid', fgColor='D6E4F0'),
            'align': _Align(horizontal='center', vertical='center'),
            'border': _B_THIN_J}
_SJ_CELL = {'font': _Font(name='微软雅黑', size=10),
            'align': _Align(horizontal='center', vertical='center'),
            'border': _B_THIN_J}
_SJ_TOTAL = {'font': _Font(name='微软雅黑', size=11, bold=True),
             'align': _Align(horizontal='center', vertical='center'),
             'border': _B_MED_TOP_J}
_MONEY = '#,##0.00'
_TITLE_FONT = _Font(name='微软雅黑', size=14, bold=True)
_PERIOD_FONT = _Font(name='微软雅黑', size=10, color='666666')
_CENTER_ALIGN = _Align(horizontal='center', vertical='center')

def _style_cell(ws, row, col, st, numfmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = st['font']
    if 'fill' in st:
        cell.fill = st['fill']
    cell.alignment = st['align']
    cell.border = st['border']
    if numfmt:
        cell.number_format = numfmt

def format_date(d):
    """Format date for readable output."""
    if d is None:
        return ''
    if isinstance(d, datetime):
        return d.strftime('%Y-%m-%d')
    if isinstance(d, date):
        return d.strftime('%Y-%m-%d')
    return str(d)

# ============================================================
# Step 2: Merchant billing files
# ============================================================
def read_analysis_LM(wb, sheet_name):
    """Read analysis data from a sheet's L-M columns (col 12-13).
    Returns list of (date, value) tuples, excluding header/total rows."""
    ws = wb[sheet_name]
    result = []
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=12).value
        val = ws.cell(row=r, column=13).value
        if label is None and val is None:
            continue
        if label in ('总计', '合计', '(空白)', '发放时间', '日期', '求和项:停', '求和项:B', None):
            continue
        # Convert date (Excel serial number, datetime, or 'YYYY-MM-DD' string)
        d = excel_serial_to_date(label)
        if d is None and isinstance(label, str):
            m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', label.strip())
            if m:
                try:
                    d = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                except ValueError:
                    d = None
        result.append((d, val if val else 0))
    return result

def read_analysis_12h(wb, floor_name):
    """Read 12h merged analysis (cols L-P). Returns dict with keys: 12h, 24h, oz"""
    ws = wb[f'{floor_name}12小时']
    result = {}
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(row=r, column=12).value  # L = 日期
        v12 = ws.cell(row=r, column=13).value or 0   # M = 12h B
        v24 = ws.cell(row=r, column=14).value or 0   # N = 24h B
        voz = ws.cell(row=r, column=15).value or 0   # O = >1天 张
        total = ws.cell(row=r, column=16).value or 0  # P = 合计
        if date_val in ('日期', '总计', '合计', None):
            continue
        d = excel_serial_to_date(date_val)
        if d is None and isinstance(date_val, str):
            m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', date_val.strip())
            if m:
                try:
                    d = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                except ValueError:
                    d = None
        result[d] = {'12h': v12, '24h': v24, 'oz': voz, 'total': total}
    return result

# ============================================================
# Header Logo Post-Processing
# ============================================================

# Logo image path (default: Desktop)
LOGO_PATH = os.path.expanduser('~/Desktop/logo.jpg')

# Header image dimensions
_HF_H_PT = 1.59 * 28.35  # 1.59 cm → ~45.08 pt
_HF_W_PT = 3.74 * 28.35  # 3.74 cm → ~106.03 pt

# Logo display size in pixels (96 dpi): 3.74 cm × 1.59 cm
_LOGO_W_PX = round(3.74 * 37.795)   # ≈ 141 px
_LOGO_H_PX = round(1.59 * 37.795)   # ≈ 60 px


def _add_logo_image(ws, logo_path=None):
    """Place the logo as a floating image anchored at the top-left (A1).

    Unlike VML header images (&G + legacyDrawingHF), floating images render
    reliably in Mac Excel / WPS / Numbers. These reports are single-page
    (fitToHeight=1), so a top-left floating image is visually equivalent to
    a header logo. Row 1 is enlarged to host the logo so it never overlaps
    the title in row 2.
    """
    if logo_path is None:
        logo_path = LOGO_PATH
    if not os.path.exists(logo_path):
        print(f"    ⚠ logo not found: {logo_path}, skipping logo image")
        return
    img = _XLImage(logo_path)
    img.width = _LOGO_W_PX
    img.height = _LOGO_H_PX
    img.anchor = 'A1'
    ws.add_image(img)
    cur = ws.row_dimensions[1].height
    ws.row_dimensions[1].height = max(cur, 48) if cur else 48

# OOXML namespaces
_NS_VML = 'urn:schemas-microsoft-com:vml'
_NS_OFFICE = 'urn:schemas-microsoft-com:office:office'
_NS_EXCEL = 'urn:schemas-microsoft-com:office:excel'
_NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_NS_REL = 'http://schemas.openxmlformats.org/package/2006/relationships'

_REL_VML = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing'
_REL_IMAGE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/image'
_CT_VML = 'application/vnd.openxmlformats-officedocument.vmlDrawing'


def _add_header_logo_to_xlsx(xlsx_path, logo_path=None):
    """Post-process an XLSX file to add header logo image to ALL sheets."""
    if logo_path is None:
        logo_path = LOGO_PATH
    if not os.path.exists(logo_path):
        print(f"    ⚠ logo not found: {logo_path}, skipping header image")
        return

    tmp_path = xlsx_path + '.tmp'

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:

            # ── Read all existing files ──
            existing = {}
            for item in zin.infolist():
                existing[item.filename] = zin.read(item.filename)

            # ── Find ALL sheet keys ──
            sheet_keys = []
            for fn in existing:
                if fn.startswith('xl/worksheets/') and fn.endswith('.xml') and '_rels' not in fn:
                    sheet_keys.append(fn)

            if not sheet_keys:
                for fn, data in existing.items():
                    zout.writestr(fn, data)
                return

            # ── Count existing VML files (one shared VML for all sheets) ──
            vml_count = sum(1 for fn in existing
                          if fn.startswith('xl/drawings/vmlDrawing') and fn.endswith('.vml'))
            vml_count += 1
            vml_fn = f'xl/drawings/vmlDrawing{vml_count}.vml'
            vml_rels_fn = f'xl/drawings/_rels/vmlDrawing{vml_count}.vml.rels'

            # ── Pre-calculate rel IDs per sheet ──
            sheet_info = {}  # sheet_key -> (next_rid, sheet_rels_key)
            for sk in sheet_keys:
                srels_key = f'xl/worksheets/_rels/{os.path.basename(sk)}.rels'
                srels_data = existing.get(srels_key)
                next_rid = 1
                if srels_data:
                    rids = re.findall(r'rId(\d+)', srels_data.decode('utf-8'))
                    if rids:
                        next_rid = max(int(x) for x in rids) + 1
                sheet_info[sk] = (next_rid, srels_key)

            # ── Process each existing file ──
            content_types_data = None
            for fn, data in existing.items():
                if fn == '[Content_Types].xml':
                    content_types_data = data
                    continue

                # Modify each sheet XML with STRING ops (preserve original XML structure)
                if fn in sheet_info:
                    txt = data.decode('utf-8')

                    # 1. Fix pageSetUpPr: MUST have fitToPage="1"
                    if '<pageSetUpPr' in txt:
                        if 'fitToPage' not in txt[txt.index('<pageSetUpPr'):txt.index('<pageSetUpPr')+60]:
                            txt = txt.replace('<pageSetUpPr', '<pageSetUpPr fitToPage="1"')
                    elif '<sheetPr>' in txt:
                        txt = txt.replace('<sheetPr>', '<sheetPr><pageSetUpPr fitToPage="1"/>')
                    else:
                        txt = txt.replace('<worksheet', '<worksheet><sheetPr><pageSetUpPr fitToPage="1"/></sheetPr>')

                    # 2. Fix pageSetup: force exact attribute order (fitToWidth before fitToHeight)
                    if '<pageSetup' in txt:
                        txt = re.sub(
                            r'<pageSetup[^/]*/>',
                            '<pageSetup paperSize="9" orientation="portrait" fitToWidth="1" fitToHeight="1"/>',
                            txt
                        )
                    else:
                        ps_tag = '<pageSetup paperSize="9" orientation="portrait" fitToWidth="1" fitToHeight="1"/>'
                        if '<pageMargins' in txt:
                            txt = txt.replace('<pageMargins', ps_tag + '<pageMargins')
                        else:
                            txt = txt.replace('</worksheet>', ps_tag + '</worksheet>')

                    # 3. Fix printOptions: ensure horizontal centering
                    if '<printOptions' in txt:
                        txt = re.sub(
                            r'<printOptions[^/]*/>',
                            '<printOptions horizontalCentered="1" verticalCentered="0"/>',
                            txt
                        )
                    else:
                        po_tag = '<printOptions horizontalCentered="1" verticalCentered="0"/>'
                        if '<pageMargins' in txt:
                            txt = txt.replace('<pageMargins', po_tag + '<pageMargins')
                        else:
                            txt = txt.replace('</worksheet>', po_tag + '</worksheet>')

                    # 4. Fix pageMargins: ensure standard margins
                    margins_xml = '<pageMargins left="0.4" right="0.4" top="0.5" bottom="0.5" header="0.3" footer="0.3"/>'
                    if '<pageMargins' in txt:
                        txt = re.sub(r'<pageMargins[^/]*/>', margins_xml, txt)
                    else:
                        txt = txt.replace('</worksheet>', margins_xml + '</worksheet>')

                    # 5. Remove old legacyDrawing / legacyDrawingHF elements
                    txt = re.sub(r'<legacyDrawingHF[^>]*/>', '', txt)
                    txt = re.sub(r'<legacyDrawing[^>]*/>', '', txt)

                    # 6. Add legacyDrawingHF before </worksheet> (VML image for header &G)
                    ld_tag = f'<legacyDrawingHF r:id="rId{sheet_info[fn][0]}"/>'
                    txt = txt.replace('</worksheet>', ld_tag + '</worksheet>')

                    data = txt.encode('utf-8')

                zout.writestr(fn, data)

            # ── Write/update sheet rels for each sheet ──
            for sk, (next_rid, srels_key) in sheet_info.items():
                vml_rel_xml = f'<Relationship Id="rId{next_rid}" Type="{_REL_VML}" Target="../drawings/vmlDrawing{vml_count}.vml"/>'
                if srels_key in existing:
                    txt = existing[srels_key].decode('utf-8')
                    txt = txt.replace('</Relationships>', vml_rel_xml + '</Relationships>')
                    zout.writestr(srels_key, txt.encode('utf-8'))
                else:
                    txt = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<Relationships xmlns="{_NS_REL}">\n {vml_rel_xml}\n</Relationships>'
                    zout.writestr(srels_key, txt.encode('utf-8'))

            # ── VML drawing (shared by all sheets) ──
            vml = f'''<xml xmlns:v="{_NS_VML}" xmlns:o="{_NS_OFFICE}" xmlns:x="{_NS_EXCEL}" xmlns:r="{_NS_R}">
 <o:shapelayout v:ext="edit">
  <o:idmap v:ext="edit" data="1"/>
 </o:shapelayout>
 <v:shapetype id="_x0000_t75" coordsize="21600,21600" o:spt="75" o:preferrelative="t"
   path="m@4@5l@4@11@9@11@9@5xe" filled="f" stroked="f">
  <v:stroke joinstyle="miter"/>
  <v:formulas>
   <v:f eqn="if lineDrawn pixelLineWidth 0"/>
   <v:f eqn="sum @0 1 0"/>
   <v:f eqn="sum 0 0 @1"/>
   <v:f eqn="prod @2 1 2"/>
   <v:f eqn="prod @3 21600 pixelWidth"/>
   <v:f eqn="prod @3 21600 pixelHeight"/>
   <v:f eqn="sum @0 0 1"/>
   <v:f eqn="prod @6 1 2"/>
   <v:f eqn="prod @7 21600 pixelWidth"/>
   <v:f eqn="prod @7 21600 pixelHeight"/>
  </v:formulas>
  <v:path o:extrusionok="f" gradientshapeok="t" o:connecttype="rect"/>
  <o:lock v:ext="edit" aspectratio="t"/>
 </v:shapetype>
 <v:shape id="_x0000_s1025" type="#_x0000_t75"
   style="position:absolute;margin-left:0;margin-top:0;width:{_HF_W_PT:.2f}pt;height:{_HF_H_PT:.2f}pt;z-index:1">
  <v:imagedata o:title="logo" r:id="rId1"/>
  <o:lock v:ext="edit" rotation="t"/>
 </v:shape>
</xml>'''
            zout.writestr(vml_fn, vml.encode('utf-8'))

            # ── VML rels ──
            vml_rels = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{_NS_REL}">
 <Relationship Id="rId1" Type="{_REL_IMAGE}" Target="../media/logo.jpg"/>
</Relationships>'''
            zout.writestr(vml_rels_fn, vml_rels.encode('utf-8'))

            # ── Logo image ──
            if 'xl/media/logo.jpg' not in existing:
                with open(logo_path, 'rb') as lf:
                    zout.writestr('xl/media/logo.jpg', lf.read())

            # ── Content types ──
            if content_types_data is not None:
                ct_txt = content_types_data.decode('utf-8')
                vml_part = f'/xl/drawings/vmlDrawing{vml_count}.vml'
                if vml_part not in ct_txt:
                    ct_txt = ct_txt.replace('</Types>',
                        f'<Override PartName="{vml_part}" ContentType="{_CT_VML}"/></Types>')
                if 'Extension="jpg"' not in ct_txt:
                    ct_txt = ct_txt.replace('</Types>',
                        '<Default Extension="jpg" ContentType="image/jpeg"/></Types>')
                zout.writestr('[Content_Types].xml', ct_txt.encode('utf-8'))

    # Replace original
    shutil.move(tmp_path, xlsx_path)


def create_merchant_sheet(ws, data, title, price, period_text=''):
    """Create a professional merchant billing sheet for client presentation."""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter

    total_quantity = 0

    # --- Styles ---
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bottom_border = Border(bottom=Side(style='thin'))
    top_bottom_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    title_font = Font(name='微软雅黑', size=16, bold=True)
    period_font = Font(name='微软雅黑', size=10)
    header_font = Font(name='微软雅黑', size=10, bold=True)
    data_font = Font(name='微软雅黑', size=10)
    total_font = Font(name='微软雅黑', size=11, bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Row 1: logo (set by _add_logo_image, height 48pt)
    # Rows 2-3: blank spacers
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10

    # --- Row 4: title (merged A4:D4) ---
    ws.merge_cells('A4:D4')
    title_cell = ws.cell(row=4, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_align
    ws.row_dimensions[4].height = 38

    # --- Row 5: period (merged A5:D5) ---
    ws.merge_cells('A5:D5')
    period_cell = ws.cell(row=5, column=1, value=period_text)
    period_cell.font = period_font
    period_cell.alignment = center_align
    ws.row_dimensions[5].height = 24

    # --- Row 6: headers ---
    headers = ['类型', '日期', '一小时（张）', f'金额（{price}元/张）']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=ci, value=h)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
    ws.row_dimensions[6].height = 32

    # --- Data rows ---
    data_start = 7
    data_end = 6 + len(data)

    for ri, (d, quantity) in enumerate(data):
        r = data_start + ri
        ws.row_dimensions[r].height = 28

        # Col B: 日期
        cell_b = ws.cell(row=r, column=2, value=format_date(d) if d else '')
        cell_b.font = data_font
        cell_b.alignment = center_align
        cell_b.border = thin_border

        # Col C: 一小时（张）
        cell_c = ws.cell(row=r, column=3, value=quantity)
        cell_c.font = data_font
        cell_c.alignment = center_align
        cell_c.border = thin_border

        # Col D: 金额
        cell_d = ws.cell(row=r, column=4, value=quantity * price)
        cell_d.font = data_font
        cell_d.alignment = center_align
        cell_d.border = thin_border
        cell_d.number_format = '#,##0.00'

        total_quantity += quantity

    # Merge A column: one big "电子劵" cell
    if data:
        ws.merge_cells(start_row=data_start, start_column=1, end_row=data_end, end_column=1)
        cell_a = ws.cell(row=data_start, column=1, value='电子劵')
        cell_a.font = data_font
        cell_a.alignment = center_align
        for rr in range(data_start, data_end + 1):
            ws.cell(row=rr, column=1).border = thin_border

    # --- Total row ---
    r = 7 + len(data)
    ws.row_dimensions[r].height = 30

    # Top border line above total
    total_border = Border(
        top=Side(style='medium'),
        bottom=Side(style='thin'),
        left=Side(style='thin'),
        right=Side(style='thin')
    )

    # Merge A+B into "合计"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell_ta = ws.cell(row=r, column=1, value='合计')
    cell_ta.font = total_font
    cell_ta.alignment = center_align
    cell_ta.border = total_border
    ws.cell(row=r, column=2).border = total_border

    cell_tc = ws.cell(row=r, column=3, value=total_quantity)
    cell_tc.font = total_font
    cell_tc.alignment = center_align
    cell_tc.border = total_border

    cell_td = ws.cell(row=r, column=4, value=total_quantity * price)
    cell_td.font = total_font
    cell_td.alignment = center_align
    cell_td.border = total_border
    cell_td.number_format = '#,##0.00'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 19

    # --- Standardized page setup (v3.5) ---
    _apply_page_setup(ws)

def step2_merchant_files(main_report_path, output_dir):
    """Generate all merchant billing files."""
    print("\n" + "=" * 60)
    print("[第二步] 生成商家计费文件")
    print("=" * 60)

    wb = openpyxl.load_workbook(main_report_path, data_only=True)

    # Determine period text from the data
    # Find the earliest and latest dates in the main sheet (scan all rows)
    ws_main = wb['停车票核销记录报表']
    min_date = None
    max_date = None
    for r in range(2, ws_main.max_row + 1):
        d = excel_serial_to_date(ws_main.cell(row=r, column=5).value)  # E = 发放时间
        if d:
            d_clean = d.date() if isinstance(d, datetime) else (d if isinstance(d, date) else None)
            if d_clean:
                if min_date is None or d_clean < min_date:
                    min_date = d_clean
                if max_date is None or d_clean > max_date:
                    max_date = d_clean

    if min_date is None:
        min_date = date.today()
    if max_date is None:
        max_date = date.today()
    period_text = f'{min_date.strftime("%Y-%m-%d")} 至 {max_date.strftime("%Y-%m-%d")}'

    for sheet_name, cfg in MERCHANT_FILES.items():
        # Read analysis data from main report
        data = read_analysis_LM(wb, sheet_name)
        if not data:
            print(f"  [{sheet_name}] 无数据，跳过")
            continue

        # Create new workbook
        out_wb = openpyxl.Workbook()
        # Remove default sheet
        out_wb.remove(out_wb.active)

        # Create the month sheet (name = ending month from period)
        m = re.search(r'至\s*(\d{4})-(\d{1,2})', period_text)
        if not m:
            m = re.search(r'(\d{4})-(\d{1,2})', period_text)
        sheet_name = f'{m.group(1)}.{int(m.group(2)):02d}' if m else 'Sheet1'
        ws = out_wb.create_sheet(title=sheet_name)
        create_merchant_sheet(ws, data, cfg['title'], cfg['price'], period_text)

        out_path = os.path.join(output_dir, cfg['filename'])
        out_wb.save(out_path)
        _add_header_logo_to_xlsx(out_path)

        print(f"  [{sheet_name}] → {cfg['filename']} ({len(data)} 天数据)")

    wb.close()
    return period_text

# ============================================================
# Step 3: 集团/九号车场电子券报表
# ============================================================
def step3_jituan_report(main_report_path, output_dir, period_text):
    """Generate 集团/九号 report."""
    print("\n" + "=" * 60)
    print("[第三步] 生成集团/九号车场电子券报表")
    print("=" * 60)

    wb = openpyxl.load_workbook(main_report_path, data_only=True)

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    # ---- 35楼集团011 ----
    ws35 = out_wb.create_sheet(title='35楼集团011')
    create_group_sheet(ws35, wb, '35楼', '011', period_text)

    # ---- 46楼集团007 ----
    ws46 = out_wb.create_sheet(title='46楼集团007')
    create_group_sheet(ws46, wb, '46楼', '007', period_text)

    # ---- 九号电子劵 ----
    ws9 = out_wb.create_sheet(title='九号电子劵')
    create_jiuhao_sheet(ws9, wb, period_text)

    out_path = os.path.join(output_dir, '集团、九号车场电子券报表.xlsx')
    out_wb.save(out_path)
    _add_header_logo_to_xlsx(out_path)
    print(f"  已生成: {out_path}")
    wb.close()

def create_group_sheet(ws, wb, floor, account, period_text):
    """Create a group sheet (35楼集团011 or 46楼集团007)."""
    # Rows 2-3: blank spacers
    # Row 4 - title
    ws.cell(row=4, column=1, value=f'集 团 {account} 账 号 免 费 停 车 数 据（电子）')
    # Row 5 - period (reference)
    ws.cell(row=5, column=1, value=period_text)
    # Row 6 - headers
    ws.cell(row=6, column=1, value='日期')
    ws.cell(row=6, column=2, value='三小时（张）')
    ws.cell(row=6, column=3, value='金额（2元/张）')
    ws.cell(row=6, column=4, value='十二小时（张）')
    ws.cell(row=6, column=5, value='金额（4元/张）')
    ws.cell(row=6, column=6, value='合计')

    # Read 3h data from L-M columns
    data_3h = read_analysis_LM(wb, f'{floor}3小时')
    # Read 12h data - we need just the 12h B values, not the combined
    # Actually for group, we need: B=3h ticket count, D=12h+24h+>1day ticket count
    # Let me re-read the structure. 12h sheet has L-P with merged data.
    # For the group report, D column = 12h + 24h + >1day ticket count

    # Read from 12h sheet analysis area (already has merged total)
    data_12h = read_analysis_12h(wb, floor)

    # Merge all dates
    all_dates = set()
    d3_map = {}
    for d, v in data_3h:
        d3_map[d] = v
        all_dates.add(d)

    # For 12h+ combined, we read from the 12h merged analysis
    d12_map = {}
    for d, vals in data_12h.items():
        # D column = 12h B total (not merged)
        # Actually for the group report format, 十二小时 = 12h+24h+>1天 total
        d12_map[d] = vals['total']  # Total = 12h+24h+>1天 combined
        all_dates.add(d)

    # Sort dates
    sorted_dates = sorted(d for d in all_dates if d is not None)

    r = 7
    total_3h = 0
    total_12h = 0
    for d in sorted_dates:
        v3 = int(d3_map.get(d, 0))
        v12 = int(d12_map.get(d, 0))

        ws.cell(row=r, column=1, value=format_date(d))
        ws.cell(row=r, column=2, value=v3)
        ws.cell(row=r, column=3, value=v3 * 2)
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=4, value=v12)
        ws.cell(row=r, column=5, value=v12 * 4)
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=6, value=v3 * 2 + v12 * 4)
        ws.cell(row=r, column=6).number_format = '#,##0.00'

        total_3h += v3
        total_12h += v12
        r += 1

    # Total row
    ws.cell(row=r, column=1, value='合计')
    ws.cell(row=r, column=2, value=total_3h)
    ws.cell(row=r, column=3, value=total_3h * 2)
    ws.cell(row=r, column=3).number_format = '#,##0.00'
    ws.cell(row=r, column=4, value=total_12h)
    ws.cell(row=r, column=5, value=total_12h * 4)
    ws.cell(row=r, column=5).number_format = '#,##0.00'
    ws.cell(row=r, column=6, value=total_3h * 2 + total_12h * 4)
    ws.cell(row=r, column=6).number_format = '#,##0.00'

    # Column widths
    for col, w in [('A', 14), ('B', 16), ('C', 16), ('D', 16), ('E', 16), ('F', 12)]:
        ws.column_dimensions[col].width = w

    # --- 排版 ---
    nc = 6
    total_r = r
    data_end = r - 1
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=nc)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=nc)
    ws.row_dimensions[4].height = 32
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 32
    for c in range(1, nc + 1):
        _style_cell(ws, 6, c, _SJ_HDR)
    for rr in range(7, data_end + 1):
        ws.row_dimensions[rr].height = 28
        for c in range(1, nc + 1):
            _style_cell(ws, rr, c, _SJ_CELL, _MONEY if c in (3, 5, 6) else None)
    ws.row_dimensions[total_r].height = 30
    for c in range(1, nc + 1):
        _style_cell(ws, total_r, c, _SJ_TOTAL, _MONEY if c in (3, 5, 6) else None)
    tc4 = ws.cell(row=4, column=1)
    tc4.font = _TITLE_FONT; tc4.alignment = _CENTER_ALIGN
    pc5 = ws.cell(row=5, column=1)
    pc5.font = _PERIOD_FONT; pc5.alignment = _CENTER_ALIGN

    # --- Standardized page setup (v3.5) ---
    _apply_page_setup(ws)

def create_jiuhao_sheet(ws, wb, period_text):
    """Create 九号电子劵 sheet."""
    # Rows 2-3: blank spacers
    # Row 4 - title
    ws.cell(row=4, column=1, value='九 号 行 馆 停 车 券 计 费 明 细 表（电子）')
    # Row 5 - period
    ws.cell(row=5, column=1, value=period_text)
    # Row 6 - headers (2 levels)
    ws.cell(row=6, column=1, value='日期')
    ws.cell(row=6, column=2, value='≤三小时（张）')
    ws.cell(row=6, column=3, value=None)
    ws.cell(row=6, column=4, value='金额（2元/张）')
    ws.cell(row=6, column=5, value='＞三小时（张）')
    ws.cell(row=6, column=6, value=None)
    ws.cell(row=6, column=7, value='金额（4元/张）')
    ws.cell(row=6, column=8, value='合计')

    # Sub-headers
    ws.cell(row=7, column=2, value='4F')
    ws.cell(row=7, column=3, value='6F')
    ws.cell(row=7, column=5, value='4F')
    ws.cell(row=7, column=6, value='6F')

    # Read data from main report
    # B (4F ≤3h) = 4楼3小时 M列
    data_4f_3h = read_analysis_LM(wb, '4楼3小时')
    # C (6F ≤3h) = 6楼3小时 M列
    data_6f_3h = read_analysis_LM(wb, '6楼3小时')
    # E (4F >3h) = 4楼12小时 合计(P)列
    data_4f_12h = read_analysis_12h(wb, '4楼')
    # F (6F >3h) = 6楼12小时 合计(P)列
    data_6f_12h = read_analysis_12h(wb, '6楼')

    # Build date maps
    d4f3_map = {d: int(v) for d, v in data_4f_3h}
    d6f3_map = {d: int(v) for d, v in data_6f_3h}
    d4f12_map = {d: int(v['total']) for d, v in data_4f_12h.items()}
    d6f12_map = {d: int(v['total']) for d, v in data_6f_12h.items()}

    all_dates = set()
    all_dates.update(d4f3_map.keys())
    all_dates.update(d6f3_map.keys())
    all_dates.update(d4f12_map.keys())
    all_dates.update(d6f12_map.keys())

    sorted_dates = sorted(d for d in all_dates if d is not None)

    r = 8
    total_b = 0; total_c = 0; total_e = 0; total_f = 0
    for d in sorted_dates:
        b = d4f3_map.get(d, 0)
        c = d6f3_map.get(d, 0)
        e = d4f12_map.get(d, 0)
        f = d6f12_map.get(d, 0)

        ws.cell(row=r, column=1, value=format_date(d))
        ws.cell(row=r, column=2, value=b)
        ws.cell(row=r, column=3, value=c)
        ws.cell(row=r, column=4, value=(b + c) * 2)
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        ws.cell(row=r, column=5, value=e)
        ws.cell(row=r, column=6, value=f)
        ws.cell(row=r, column=7, value=(e + f) * 4)
        ws.cell(row=r, column=7).number_format = '#,##0.00'
        ws.cell(row=r, column=8, value=(b + c) * 2 + (e + f) * 4)
        ws.cell(row=r, column=8).number_format = '#,##0.00'

        total_b += b; total_c += c; total_e += e; total_f += f
        r += 1

    # Total row
    ws.cell(row=r, column=1, value='合计')
    ws.cell(row=r, column=2, value=total_b)
    ws.cell(row=r, column=3, value=total_c)
    ws.cell(row=r, column=4, value=(total_b + total_c) * 2)
    ws.cell(row=r, column=4).number_format = '#,##0.00'
    ws.cell(row=r, column=5, value=total_e)
    ws.cell(row=r, column=6, value=total_f)
    ws.cell(row=r, column=7, value=(total_e + total_f) * 4)
    ws.cell(row=r, column=7).number_format = '#,##0.00'
    ws.cell(row=r, column=8, value=(total_b + total_c) * 2 + (total_e + total_f) * 4)
    ws.cell(row=r, column=8).number_format = '#,##0.00'

    # Column widths
    for col, w in [('A', 14), ('B', 6), ('C', 6), ('D', 16), ('E', 6), ('F', 6), ('G', 16), ('H', 12)]:
        ws.column_dimensions[col].width = w

    # --- 排版 ---
    nc = 8
    total_r = r
    data_end = r - 1
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=nc)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=nc)
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=3)
    ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=6)
    ws.row_dimensions[4].height = 32
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 32
    ws.row_dimensions[7].height = 28
    for c in range(1, nc + 1):
        _style_cell(ws, 6, c, _SJ_HDR)
        _style_cell(ws, 7, c, _SJ_HDR)
    for rr in range(8, data_end + 1):
        ws.row_dimensions[rr].height = 28
        for c in range(1, nc + 1):
            _style_cell(ws, rr, c, _SJ_CELL, _MONEY if c in (4, 7, 8) else None)
    ws.row_dimensions[total_r].height = 30
    for c in range(1, nc + 1):
        _style_cell(ws, total_r, c, _SJ_TOTAL, _MONEY if c in (4, 7, 8) else None)
    tc4 = ws.cell(row=4, column=1)
    tc4.font = _TITLE_FONT; tc4.alignment = _CENTER_ALIGN
    pc5 = ws.cell(row=5, column=1)
    pc5.font = _PERIOD_FONT; pc5.alignment = _CENTER_ALIGN

    # --- Standardized page setup (v3.5) ---
    _apply_page_setup(ws)

# ============================================================
# Step 4: 车场报表 (来访停车卡)
# ============================================================
def step4_chechang_report(main_report_path, vehicle_report_path, output_dir):
    """Generate 车场报表."""
    print("\n" + "=" * 60)
    print("[第四步] 生成车场报表(来访停车卡)")
    print("=" * 60)

    wb_main = openpyxl.load_workbook(main_report_path, data_only=True)
    wb_veh = openpyxl.load_workbook(vehicle_report_path, data_only=True)
    ws_veh = wb_veh['车辆进出报表']

    # Build vehicle entry/exit lookup: (plate, entry_date) -> (entry_date, entry_time, exit_date, exit_time)
    veh_data = {}  # key: (plate, date_key)
    for r in range(2, ws_veh.max_row + 1):
        plate = ws_veh.cell(row=r, column=2).value
        if not plate:
            continue
        # Entry
        entry_date_val = ws_veh.cell(row=r, column=3).value  # date part
        entry_time_val = ws_veh.cell(row=r, column=4).value  # time part
        # Exit
        exit_date_val = ws_veh.cell(row=r, column=5).value
        exit_time_val = ws_veh.cell(row=r, column=6).value

        # Normalize plate
        plate = str(plate).strip()

        # Get date key from entry
        entry_date_key = parse_date_value(entry_date_val)
        if not entry_date_key:
            continue

        key = (plate, entry_date_key)
        if key not in veh_data:
            veh_data[key] = {
                'entry_date': entry_date_val,
                'entry_time': entry_time_val,
                'exit_date': exit_date_val,
                'exit_time': exit_time_val,
            }

    # Define sheet mappings
    sheet_mappings = [
        ('物业', '天启来访停车卡', '集团来访停车卡明细表'),
        ('哥弟', '附件七(哥弟公司来方停车卡)', '哥弟公司来访停车卡明细表'),
        ('商管', '商业公司来访停车卡', '商业公司来访停车卡明细表'),
    ]

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    for sub_sheet, out_sheet_name, title in sheet_mappings:
        ws = out_wb.create_sheet(title=out_sheet_name)
        has_data = create_parking_card_sheet(ws, wb_main, veh_data, sub_sheet, title)
        if not has_data:
            out_wb.remove(ws)
            print(f"  [{sub_sheet}] 无数据，跳过")
        else:
            print(f"  [{sub_sheet}] → {out_sheet_name}")

    out_path = os.path.join(output_dir, '车场报表(物业、集团、哥弟来访条).xlsx')
    if not out_wb.sheetnames:
        print("  ⚠️ 所有子表均无数据，跳过生成车场报表")
        wb_main.close()
        wb_veh.close()
        return
    out_wb.save(out_path)
    _add_header_logo_to_xlsx(out_path)
    print(f"  已生成: {out_path}")

    wb_main.close()
    wb_veh.close()

def create_parking_card_sheet(ws, wb_main, veh_data, sub_sheet, title):
    """Create a parking card sheet."""
    # Rows 2-3: blank spacers
    # Row 4 - title
    ws.cell(row=4, column=1, value=title)
    # Read data from main report sub-sheet
    ws_sub = wb_main[sub_sheet]

    # Extract plate + date from sub-sheet data rows (columns G and E)
    records = []
    for r in range(2, ws_sub.max_row + 1):
        plate = ws_sub.cell(row=r, column=7).value
        issue_time = ws_sub.cell(row=r, column=5).value
        if plate and issue_time:
            plate = str(plate).strip()
            issue_date = parse_date_value(issue_time)
            if not issue_date:
                continue
            records.append((plate, issue_date, issue_time))

    if not records:
        return False

    # Determine the period date (latest entry)
    all_dates = [rec[1] for rec in records if rec[1]]
    period_date = max(all_dates) if all_dates else datetime.now().date()

    # Row 5 - period
    if isinstance(period_date, date):
        ws.cell(row=5, column=1, value=f'{period_date.year}年{period_date.month}月')

    # Row 6 - headers
    ws.cell(row=6, column=1, value='序号')
    ws.cell(row=6, column=2, value='进场日期')
    ws.cell(row=6, column=3, value='进场时间')
    ws.cell(row=6, column=4, value='出场日期')
    ws.cell(row=6, column=5, value='出场时间')
    ws.cell(row=6, column=6, value='车牌号码')
    ws.cell(row=6, column=7, value='免费额')
    ws.cell(row=6, column=8, value='备注')

    # Match each record with vehicle data
    r = 7
    seq = 1
    total_free = 0

    for plate, issue_date, issue_time in records:
        # Find matching vehicle entry
        key = (plate, issue_date)
        veh = veh_data.get(key)

        if veh is None:
            # Try other dates nearby (within a day)
            for offset in [-1, 1]:
                alt_date = issue_date + timedelta(days=offset)
                veh = veh_data.get((plate, alt_date))
                if veh:
                    break

        ws.cell(row=r, column=1, value=seq)

        # Entry date + time
        if veh:
            ed_val = veh['entry_date']
            if isinstance(ed_val, datetime):
                ws.cell(row=r, column=2, value=ed_val.strftime('%Y-%m-%d'))
            else:
                ws.cell(row=r, column=2, value=str(ed_val) if ed_val else None)
            ws.cell(row=r, column=3, value=veh['entry_time'])

            # Exit date + time
            exit_date = veh['exit_date']
            exit_time = veh['exit_time']

            if exit_date and str(exit_date) not in ('null', 'None'):
                if isinstance(exit_date, datetime):
                    ws.cell(row=r, column=4, value=exit_date.strftime('%Y-%m-%d'))
                else:
                    ws.cell(row=r, column=4, value=str(exit_date))
            if exit_time and str(exit_time) not in ('null', 'None'):
                ws.cell(row=r, column=5, value=exit_time)
        else:
            # No match found - show issue date only
            if isinstance(issue_date, datetime):
                ws.cell(row=r, column=2, value=issue_date.strftime('%Y-%m-%d'))
            elif isinstance(issue_date, date):
                ws.cell(row=r, column=2, value=issue_date.strftime('%Y-%m-%d'))
            else:
                ws.cell(row=r, column=2, value=str(issue_date) if issue_date else None)

        ws.cell(row=r, column=6, value=plate)

        # Free amount: only calculate if exit data exists AND same day (no cross-day)
        if veh and veh['exit_date'] and str(veh['exit_date']) not in ('null', 'None'):
            ed = veh['entry_date']
            et = veh['entry_time']
            xd = veh['exit_date']
            xt = veh['exit_time']

            # Check if dates are valid
            ed_ok = ed and str(ed) not in ('null', 'None')
            xd_ok = xd and str(xd) not in ('null', 'None')
            et_ok = et and str(et) not in ('null', 'None')
            xt_ok = xt and str(xt) not in ('null', 'None')

            if ed_ok and xd_ok and et_ok and xt_ok:
                # Check if same day
                if isinstance(ed, datetime):
                    ed_d = ed.date()
                elif isinstance(ed, date):
                    ed_d = ed
                else:
                    ed_d = None

                if isinstance(xd, datetime):
                    xd_d = xd.date()
                elif isinstance(xd, date):
                    xd_d = xd
                else:
                    xd_d = None

                if ed_d and xd_d and ed_d == xd_d:
                    # Same day - use formula
                    try:
                        # Calculate hours difference
                        if isinstance(et, datetime):
                            et_t = et.time()
                        else:
                            et_t = et
                        if isinstance(xt, datetime):
                            xt_t = xt.time()
                        else:
                            xt_t = xt

                        # Convert time to hours
                        if hasattr(et_t, 'hour'):
                            et_h = et_t.hour + et_t.minute / 60 + et_t.second / 3600
                            xt_h = xt_t.hour + xt_t.minute / 60 + xt_t.second / 3600
                        else:
                            # time fraction (Excel serial)
                            et_h = float(et) * 24 if et else 0
                            xt_h = float(xt) * 24 if xt else 0

                        diff_h = xt_h - et_h
                        if diff_h < 0:
                            diff_h += 24  # overnight but same calendar day? rare

                        hours = int(diff_h)
                        free = (hours + 1) * 4
                        ws.cell(row=r, column=7, value=free)
                        ws.cell(row=r, column=7).number_format = '#,##0.00'
                        total_free += free
                    except:
                        pass  # Leave blank if calculation fails
                # else: cross-day, leave blank for manual
        # else: no exit data, leave blank

        seq += 1
        r += 1

    # Total row
    ws.cell(row=r, column=1, value='合计')
    ws.cell(row=r, column=7, value=total_free)
    ws.cell(row=r, column=7).number_format = '#,##0.00'

    # Column widths
    for col, w in [('A', 8), ('B', 14), ('C', 12), ('D', 14), ('E', 12), ('F', 14), ('G', 10), ('H', 10)]:
        ws.column_dimensions[col].width = w

    # --- 排版 ---
    nc = 8
    total_r = r
    data_end = r - 1
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=nc)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=nc)
    ws.row_dimensions[4].height = 32
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 32
    for c in range(1, nc + 1):
        _style_cell(ws, 6, c, _SJ_HDR)
    for rr in range(7, data_end + 1):
        ws.row_dimensions[rr].height = 28
        for c in range(1, nc + 1):
            _style_cell(ws, rr, c, _SJ_CELL, _MONEY if c == 7 else None)
    ws.row_dimensions[total_r].height = 30
    for c in range(1, nc + 1):
        _style_cell(ws, total_r, c, _SJ_TOTAL, _MONEY if c == 7 else None)
    tc4 = ws.cell(row=4, column=1)
    tc4.font = _TITLE_FONT; tc4.alignment = _CENTER_ALIGN
    pc5 = ws.cell(row=5, column=1)
    pc5.font = _PERIOD_FONT; pc5.alignment = _CENTER_ALIGN

    # --- Standardized page setup (v3.5) ---
    _apply_page_setup(ws)

    return True

# ============================================================
# Main
# ============================================================
def main():
    global MAIN_REPORT, VEHICLE_REPORT, OUTPUT_DIR

    if len(sys.argv) < 4:
        print("用法: python3 generate_branch.py <主报表.xlsx> <车辆进出报表.xlsx> <输出目录>")
        sys.exit(1)

    MAIN_REPORT = os.path.expanduser(sys.argv[1])
    VEHICLE_REPORT = os.path.expanduser(sys.argv[2])
    OUTPUT_DIR = os.path.expanduser(sys.argv[3])

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("=" * 60)
    print("停车票分支报表生成脚本")
    print(f"主报表: {MAIN_REPORT}")
    print(f"车辆进出报表: {VEHICLE_REPORT}")
    print(f"输出目录: {OUTPUT_DIR}")
    print("=" * 60)

    # Step 2: Merchant billing files
    period_text = step2_merchant_files(MAIN_REPORT, OUTPUT_DIR)

    # Step 3: 集团/九号 报表
    step3_jituan_report(MAIN_REPORT, OUTPUT_DIR, period_text)

    # Step 4: 车场报表
    step4_chechang_report(MAIN_REPORT, VEHICLE_REPORT, OUTPUT_DIR)

    print("\n" + "=" * 60)
    print("✅ 全部完成!")
    print("=" * 60)

if __name__ == '__main__':
    main()
