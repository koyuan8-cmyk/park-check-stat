#!/usr/bin/env python3
"""
停车票核销报表 → HTML 生成器
用法: python3 generate_html.py <主报表.xlsx> [输出.html]
"""

import openpyxl
import sys, os
from datetime import datetime, date
from html import escape

SHEET_ORDER = [
    '停车票核销记录报表',
    '爱康', '爱康口腔', '爱来', '阿玛施', '歌蕊', '皆大欢洗', '新奥美嘉',
    '丝芭', '哥弟', '精应', '物业', '商管', '集团',
    '35楼3小时', '35楼12小时', '35楼24小时',
    '46楼3小时', '46楼12小时', '46楼24小时',
    '九号',
    '九号6楼', '6楼3小时', '6楼12小时', '6楼24小时', '6楼大于一天',
    '九号4楼', '4楼3小时', '4楼12小时', '4楼24小时', '4楼大于一天',
]

CSS = """
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Microsoft YaHei','PingFang SC',sans-serif;background:#f0f2f5;color:#333}
.header{background:#1a1a2e;color:#fff;padding:18px 24px;position:sticky;top:0;z-index:100;box-shadow:0 2px 8px rgba(0,0,0,0.15)}
.header h1{font-size:20px;font-weight:600}
.header .meta{font-size:12px;color:#aab;margin-top:4px}
.tabs{display:flex;flex-wrap:wrap;gap:4px;padding:12px 16px;background:#fff;border-bottom:1px solid #e0e0e0;position:sticky;top:72px;z-index:99}
.tab{padding:6px 14px;font-size:12px;border-radius:4px;cursor:pointer;border:1px solid #ddd;background:#fafafa;white-space:nowrap;transition:all .15s}
.tab:hover{background:#e8f0fe;border-color:#4a90d9}
.tab.active{background:#1a73e8;color:#fff;border-color:#1a73e8}
.tab .count{font-size:10px;opacity:.7;margin-left:3px}
.content{padding:16px 20px}
.sheet{display:none}
.sheet.active{display:block}
.table-wrap{overflow-x:auto;background:#fff;border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,0.08)}
table{border-collapse:collapse;width:100%;font-size:12px;table-layout:auto}
th,td{padding:3px 6px;border:1px solid #e0e0e0;text-align:center;white-space:nowrap}
th{background:#f5f6f8;font-weight:600;color:#444;position:sticky;top:0;z-index:2}
tr:nth-child(even){background:#fafbfc}
tr:hover{background:#e8f0fe}
.analysis-label{background:#fff3e0!important;font-weight:600}
.analysis-header{background:#ffe0b2!important}
.blank-row td{color:#ccc;font-style:italic}
.total-row td{font-weight:700;background:#e8f5e9!important;border-top:2px solid #4caf50}
.empty-sheet{padding:40px;text-align:center;color:#999;font-size:14px}
.footer{text-align:center;padding:16px;color:#999;font-size:11px}
"""

def fmt(v):
    """Format a cell value for HTML display."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return f'{v:.4g}'
    return str(v)

def is_total_row(row_values):
    """Check if this is a total/summary row."""
    texts = [str(v).strip() for v in row_values if v is not None]
    return any(t in ('总计', '合计') for t in texts)

def is_blank_row(row_values):
    return all(v is None or str(v).strip() == '' for v in row_values)

def is_analysis_header(row_values):
    texts = [str(v).strip() for v in row_values if v is not None]
    return any(t in ('发放时间', '日期', '求和项:停', '求和项:B', '求和项:张') for t in texts)

def render_sheet(name, ws):
    """Render a single sheet as HTML table."""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))

    if not rows:
        return f'<div class="empty-sheet">暂无数据</div>'

    # Count non-empty rows
    non_empty = sum(1 for r in rows if any(v is not None for v in r))

    html = '<div class="table-wrap"><table>'

    for ri, row in enumerate(rows):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        row_class = ''
        if ri == 0:
            row_class = ' class="header-row"'
        elif is_total_row(row):
            row_class = ' class="total-row"'
        elif is_blank_row(row):
            row_class = ' class="blank-row"'

        html += f'<tr{row_class}>'

        # Determine if this is analysis section (cols L+)
        for ci, val in enumerate(row):
            cell_class = ''
            if ci >= 11 and ri > 0:
                if is_analysis_header([val]):
                    cell_class = ' class="analysis-header"'

            formatted = fmt(val)
            html += f'<td{cell_class}>{escape(formatted)}</td>'

        html += '</tr>'

    html += '</table></div>'
    return html, non_empty

def main():
    if len(sys.argv) < 2:
        print("用法: python3 generate_html.py <主报表.xlsx> [输出.html]")
        sys.exit(1)

    xlsx_path = os.path.expanduser(sys.argv[1])
    if len(sys.argv) >= 3:
        html_path = os.path.expanduser(sys.argv[2])
    else:
        base = os.path.splitext(os.path.basename(xlsx_path))[0]
        html_path = os.path.join(os.path.dirname(xlsx_path) or '.', f'{base}.html')

    print(f"📖 读取: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Collect sheet data
    sheets = []
    for name in SHEET_ORDER:
        if name in wb.sheetnames:
            ws = wb[name]
            html_fragment, count = render_sheet(name, ws)
            sheets.append((name, html_fragment, count))

    wb.close()

    # Build HTML
    tabs_html = ''
    sheets_html = ''
    for i, (name, fragment, count) in enumerate(sheets):
        active = 'active' if i == 0 else ''
        tabs_html += f'<div class="tab {active}" onclick="switchTab({i})">{escape(name)}<span class="count">({count})</span></div>\n'
        sheets_html += f'<div class="sheet {active}" id="sheet-{i}">\n{fragment}\n</div>\n'

    total_sheets = len(sheets)

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>停车票核销记录报表</title>
<style>
{CSS}
</style>
</head>
<body>

<div class="header">
  <h1>🅿️ 停车票核销记录报表</h1>
  <div class="meta">共 {total_sheets} 个子表 · 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
</div>

<div class="tabs" id="tabs">
{tabs_html}
</div>

<div class="content">
{sheets_html}
</div>

<div class="footer">停车票核销报表生成器 v3.0</div>

<script>
function switchTab(idx) {{
  document.querySelectorAll('.tab').forEach((t,i) => t.classList.toggle('active', i===idx));
  document.querySelectorAll('.sheet').forEach((s,i) => s.classList.toggle('active', i===idx));
  localStorage.setItem('parkingActiveTab', idx);
}}
// Restore last tab
(function() {{
  const saved = localStorage.getItem('parkingActiveTab');
  if (saved) switchTab(parseInt(saved));
}})();
</script>

</body>
</html>'''

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)

    size_kb = os.path.getsize(html_path) / 1024
    print(f"✅ HTML 已生成: {html_path} ({size_kb:.0f} KB)")

if __name__ == '__main__':
    main()
